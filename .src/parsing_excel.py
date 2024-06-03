import pandas as pd
import os
import json
import openpyxl
import shutil
import time
import re

ALL_EXCEL_AS_TEXT_PATH='./excel_as_text/'
ALL_EXCEL_PATH='./excel/'
METADATA_FILE_PATH='./EXCEL_METADATA.json'
step_time=0

def parsing_format(x):
    if hasattr(x, '__dict__'):
        res = {'objectType': str(type(x).__name__), 'measure': {}}
        res_name=[]
        for k, v in vars(x).items():
            if hasattr(v, '__dict__'):
                format_name, format=parsing_format(v)
                if format['measure']: 
                    res['measure'][k] = format
                    res_name.append(format_name)
            else:
                if v:
                    res['measure'][k] = v 
                    res_name.append(f"{k}={str(v)}")
        return '  '.join(res_name), res
    else:
        return x, {'objectType': str(type(x).__name__), 'measure': x}
    

def log_with_timer(msg):
    global step_time
    current_time=time.time()
    if step_time==0: step_time=current_time
    print(f'(after {"{:.3f}".format(current_time-step_time)}) {msg}')
    step_time=current_time


class SheetExcelObject:
    def __init__(self, sheet_openpyxl, excelpath: str, text_path: str, sheet_name: str, keys: list = None, header_line: int = None) -> None:
        self.sheet_openpyxl=sheet_openpyxl
        self.excelpath=excelpath
        self.text_path=text_path
        self.sheet_name=sheet_name
        self.keys=keys
        self.header_line=header_line

    def read_value(self):
        try:
            df_values=pd.read_excel(self.excelpath, sheet_name=self.sheet_name, index_col=False, header=None)
        except ValueError:
            print(f'Not found excel file [{self.sheet_name}] or sheet {self.sheet_name}')
            return None
        self.len_active_col=len(list(df_values.columns))
        self.len_active_row=len(df_values)
        self.df_values=df_values.rename(columns=dict(zip(list(df_values.columns), ALPHABET_COL_NAME[:self.len_active_col])))

    def read_style(self):
        layout_data = []
        general_format={'width': {}, 'font':{}, 'border':{} , 'fill':{}, 'number_format': {}, 'protection':{}, 'alignment':{}}
        for col_name in ALPHABET_COL_NAME[:self.len_active_col]:
            general_format['width'][col_name] = self.sheet_openpyxl.column_dimensions[col_name].width
        for row in self.sheet_openpyxl.iter_rows():
            row_layout_data = []
            for cell in row:
                cell_layout_data = {
                    'font': parsing_format(cell.font),
                    'border': parsing_format(cell.border),
                    'fill': parsing_format(cell.fill),
                    'number_format': parsing_format(cell.number_format),
                    'protection': parsing_format(cell.protection),
                    'alignment': parsing_format(cell.alignment)
                }

                for format_type, name_format in cell_layout_data.items():
                    name, format = name_format
                    cell_layout_data[format_type]=name
                    format['objectType']=format_type
                    if name not in general_format[format_type].keys():
                        general_format[format_type][name]=format

                row_layout_data.append(cell_layout_data)
            layout_data.append(row_layout_data)

        self.layout_data=layout_data
        self.general_format=general_format


    def get_record_title(self):
        row_title_list=[]
        row_title_list=[f"L{i+1}" for i in range(0, self.header_line-1)]
        if self.header_line>=1: 
            row_title_list.append("HEADER")

        if self.keys:
            if self.header_line>=1: 
                self.df_values=self.df_values.rename(columns=dict(zip(list(self.df_values.columns), self.df_values.iloc[self.header_line-1])))
            key_df=self.df_values[self.keys][self.df_values.index>=self.header_line]
            for key in self.keys:
                key_df.loc[:, key] = key_df[key].astype('str')
                key_df.loc[:, key] = key_df[key].replace(to_replace=r'\W+', value='', regex=True)
            record_title_df=key_df[self.keys].agg(' & '.join, axis=1)
            duplicate_mask = record_title_df.duplicated(keep=False)
            record_title_df[duplicate_mask] = 'L'+(record_title_df[duplicate_mask].index+1).astype(str) +' '+ record_title_df[duplicate_mask]
            row_title_list.extend(list(record_title_df))
        else:
            row_title_list.extend([f"L{i+1}" for i in range(self.header_line, self.len_active_row)])         
        return row_title_list


    def write_to_text(self):
        row_title_list=self.get_record_title()
        for i in range(self.len_active_row):
            row_title=row_title_list[i]

            record_path=f"{self.text_path}{self.sheet_name}/{row_title}/"
            record_data_path=f"{record_path}values.csv"   
            record_layout_path=f"{record_path}styles.json"

            os.makedirs(record_path)
            self.df_values.iloc[i].T.to_csv(record_data_path)
            
            with open(record_layout_path, 'w') as wf:
                json.dump(dict(zip(self.df_values.columns, self.layout_data[i])), wf, indent=2)
        
        try:
            with open(f"{self.text_path}{self.sheet_name}/styles_detail.json", 'w') as wf:
                    json.dump(self.general_format, wf, indent=2)
        except FileNotFoundError:
            pass


def gen_excel_as_text(excel_path, text_path, excel_cf, local_ALPHABET_COL_NAME):
    global ALPHABET_COL_NAME
    ALPHABET_COL_NAME = local_ALPHABET_COL_NAME
    workbook = openpyxl.load_workbook(excel_path)

    try:
        shutil.rmtree(text_path)
    except:
        pass
    
    for sheet_name in workbook.sheetnames:
        sheet_keys=None
        sheet_header_line=0
        try:
            sheet_properties=excel_cf[sheet_name]
            try:
                sheet_keys=sheet_properties['KEYS']
            except:
                pass
            try:
                sheet_header_line=sheet_properties['HEADER_LINE']
            except:
                pass
        except:
            pass

        sheet_obj=SheetExcelObject(workbook[sheet_name], excel_path, text_path, sheet_name, sheet_keys, sheet_header_line)
        #GET DATA
        log_with_timer(f"{sheet_name}: reading data...")
        sheet_obj.read_value()

        #GET LAYOUT
        log_with_timer(f"{sheet_name}: reading style...")
        sheet_obj.read_style()

        #WRITE TO READABLE FILES
        log_with_timer(f"{sheet_name}: writing data...")
        sheet_obj.write_to_text()


    
    