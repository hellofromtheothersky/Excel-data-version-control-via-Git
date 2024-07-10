import pandas as pd
import os
import json
import openpyxl
from gitexcel import py_git, log


def construct_format_object(format_dict):
    measures={}
    if isinstance(format_dict['measure'], dict):
        for k, v in format_dict['measure'].items():
            if isinstance(v, dict):
                measures[k]=construct_format_object(v)
            else:
                measures[k]=v
    else:
        measure=format_dict['measure']
            
    if format_dict['objectType']=='StyleProxy': obj=openpyxl.styles.proxy.StyleProxy(**measures)
    elif format_dict['objectType']=='font': obj=openpyxl.styles.fonts.Font(**measures)
    elif format_dict['objectType']=='border': obj=openpyxl.styles.borders.Border(**measures)
    elif format_dict['objectType']=='fill': obj=openpyxl.styles.fills.PatternFill(**measures)
    elif format_dict['objectType']=='number_format': obj=measure
    elif format_dict['objectType']=='protection': obj=openpyxl.styles.protection.Protection(**measures)
    elif format_dict['objectType']=='alignment': obj=openpyxl.styles.alignment.Alignment(**measures)
    elif format_dict['objectType']=='Side': obj=openpyxl.styles.borders.Side(**measures)
    elif format_dict['objectType']=='Color': obj=openpyxl.styles.colors.Color(**measures)
    #optimize above code hmmmmmmmmmmmmm
    return obj    


def get_changed_styles_rows(sheet_path):
    # Run git status command
    new_files, modified_files, _ = py_git.get_changed_files(sheet_path)

    modified_files.extend(new_files)
    return [x for x in modified_files if x.endswith('/styles.json')]


def gather_data_files(sheet_path, change_style_files):
    sheet_values=pd.DataFrame()
    rows_format={}
    for item in os.listdir(sheet_path):
        item_path = os.path.join(sheet_path, item)
        if os.path.isdir(item_path):
            df=pd.read_csv(item_path+'/values.csv', index_col=0)
            record_order=int(df.columns[0])
            if record_order in rows_format.keys():
                df=df.rename(columns={df.columns[0]: record_order+1})
                record_order+=1
            df=df.reset_index(drop=True).T
            sheet_values=pd.concat([sheet_values, df])
            
            if item_path.strip('./')+'/styles.json' in change_style_files:
                with open(item_path+'/styles.json') as wf:
                    rows_format[record_order]=json.load(wf)
    
    sheet_values.index=sheet_values.index.astype('int')
    sheet_values=sheet_values.sort_index()
    return sheet_values, rows_format


def gen_excel_from_text(excel_path, excel_text_path, ALPHABET_COL_NAME):
    first_sheet=True
    for sheet_name in [f for f in os.listdir(excel_text_path) if os.path.isdir(os.path.join(excel_text_path, f))]:
        sheet_path=excel_text_path+'/'+sheet_name
        sheet_values=pd.DataFrame()
        rows_format={}
        change_style_files=get_changed_styles_rows(sheet_path)  
        log.print_log_info(f'Detect {sheet_path} have changed style row files {change_style_files}')      
        sheet_values, rows_format = gather_data_files(sheet_path, change_style_files)
        sheet_values=sheet_values.rename(columns=dict(zip(list(sheet_values.columns), ALPHABET_COL_NAME[:len(list(sheet_values.columns))])))
                
        #WRITE STYLES
        num_col=len(sheet_values.columns)
        workbook = openpyxl.load_workbook(excel_path)

        try:
            destination_sheet = workbook[sheet_name]
        except KeyError:
            if first_sheet:
                destination_sheet = workbook['Sheet1']
                destination_sheet.title = sheet_name
                first_sheet=False
            else:
                workbook.create_sheet(sheet_name)
                destination_sheet = workbook[sheet_name]


        with open(f"{excel_text_path}/{sheet_name}/styles_detail.json", 'r') as rf:
            sheet_style_collections=json.load(rf)
        
        for col, width in sheet_style_collections['width'].items():
            destination_sheet.column_dimensions[col].width = width

        for report_order, format in rows_format.items():
            for source_col_name, abc_col_name in list(zip(format.keys(), ALPHABET_COL_NAME[:num_col])):
                destination_cell = destination_sheet[abc_col_name+str(report_order+1)]
                for type in ['font', 'border', 'fill', 'number_format', 'protection', 'alignment']:
                    format_name=format[source_col_name][type]
                    format_object=construct_format_object(sheet_style_collections[type][format_name])
                    destination_cell.__setattr__(type, format_object)

        for irow in range(len(sheet_values)):
            for abc_col_name in ALPHABET_COL_NAME[:num_col]:
                destination_cell = destination_sheet[abc_col_name+str(irow+1)]
                destination_cell.__setattr__('value', sheet_values.iloc[irow][abc_col_name])

        workbook.save(excel_path)
        workbook.close()


